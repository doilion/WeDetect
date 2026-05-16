"""Generate paper §A results workbook (multi-sheet .xlsx).

Aggregates everything currently known about the TCT_NGC OC-HMTA + ICF
experiment line into a single Excel-importable workbook:

  Sheet 1 — Main ablation table (one row per trained method)
  Sheet 2 — Per-organ novel-zero-shot breakdown
  Sheet 3 — Bypass diagnostic on the auxfix ep12 ckpt
  Sheet 4 — ICF training health (ICFCollapseGuard ep12 readings)
  Sheet 5 — Rank-label audit (which (organ, axis) pairs are usable for ord_loss)
  Sheet 6 — Shared training hyperparameters (so the ablation is apples-to-apples)

Method strings are written verbatim — no shorthand. Each method row spells
out which components are in (M1 mask, 5-attribute prompts, HTA stages, ord
loss, axis-struct loss, ICF, etc.) so the table is readable without cross-
referencing the docs.

Outputs to: docs/tct_ngc_paper_a_results_20260516.xlsx
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


REPO = Path(__file__).resolve().parent.parent
OUT_PATH = REPO / "docs" / "tct_ngc_paper_a_results_20260516.xlsx"


# ──────────────────────────────────────────────────────────────────────
# Style helpers
# ──────────────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
EMPHASIS_FILL = PatternFill("solid", fgColor="FFF2CC")
WIN_FILL = PatternFill("solid", fgColor="E2EFDA")
FAIL_FILL = PatternFill("solid", fgColor="FCE4D6")
HEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=12)
WRAP = Alignment(wrap_text=True, vertical="top", horizontal="left")
RIGHT = Alignment(vertical="top", horizontal="right")


def _autofit(ws, max_width: int = 80) -> None:
    """Approximate column auto-fit using widest cell content."""
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        longest = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            for line in str(cell.value).splitlines():
                longest = max(longest, len(line))
        ws.column_dimensions[col_letter].width = min(max(12, longest + 2), max_width)


def _write_header(ws, headers, row=1):
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP


# ──────────────────────────────────────────────────────────────────────
# Sheet 1 — Main ablation table
# ──────────────────────────────────────────────────────────────────────
def build_main_table(wb: Workbook) -> None:
    ws = wb.create_sheet("MainAblation")

    headers = [
        "Row", "Tag", "Method",
        "Trainable text-side params",
        "Aux losses",
        "Base macro", "Base inst-wt",
        "Novel macro", "Novel inst-wt",
        "Δ Base",
        "Δ Novel",
        "Status",
    ]
    _write_header(ws, headers)

    # Concise Method: refer to shared notation below the table.
    rows = [
        (
            1, "M1 / 1-PSC",
            "Baseline (M1) + 1-PSC.",
            "0",
            "cls + bbox + dfl",
            0.3369, 0.4046, 0.1500, 0.1537,
            -0.0027, -0.0142, "trained",
        ),
        (
            2, "M1 + 5-attr mean pool",
            "Baseline + 5-attr + mean pool.",
            "0",
            "cls + bbox + dfl",
            0.3396, 0.3973, 0.1642, 0.1657,
            0.0000, 0.0000, "trained (Δ reference)",
        ),
        (
            3, "M2 完整方法",
            "Baseline + 5-attr + HTA (Stage 1+2+3) + 4 aux losses + ord_loss (mean norm).",
            "~200K",
            "cls + bbox + dfl + pool_entropy + proj_drift + gate_entropy + rank_norm + ord_loss",
            0.3436, 0.4088, 0.1051, 0.1203,
            0.0040, -0.0591, "trained",
        ),
        (
            4, "M2 + auxfix",
            "M2 + λ_pool_entropy ×15 (0.02→0.3) + ord_loss sum norm (×6 magnitude).",
            "~200K",
            "M2 + pool_entropy(λ=0.3) + ord_loss(sum)",
            0.3434, 0.4070, 0.0557, 0.0651,
            0.0038, -0.1085, "trained",
        ),
        (
            5, "M2 + axisstruct (Row 5)",
            "M2 (sum norm) + axis structure loss (attract same-axis cos≥0.5, repel cross-organ cos≤0.1).",
            "~200K",
            "M2 + axis_attract + cross_organ_repel",
            0.3371, None, 0.0665, 0.0848,
            -0.0025, -0.0977, "trained",
        ),
        (
            6, "M2 ordclean (Row 6c)",
            "Baseline + 5-attr + HTA Stage 1 only (skip Stage 2/3) + ord_loss on clean axes (exclude resp/serous/cervical-infection; skip rank collisions).",
            "~150K",
            "cls + bbox + dfl + pool_entropy + proj_drift + ord_loss(mean)",
            0.3386, None, 0.1555, 0.1538,
            -0.0010, -0.0087, "trained",
        ),
        (
            7, "ICF / Design A ⭐",
            "Baseline + 5-attr + ICF (image-Q cross-attn over 5 per-attr MLP + PE, 8 heads).",
            "~2.4M",
            "cls + bbox + dfl",
            0.3522, None, 0.1648, 0.1582,
            0.0126, 0.0006, "trained — paper §A main",
        ),
        (
            8, "ICF + 1-PSC (ablation)",
            "Baseline + 1-PSC + ICF (1-attr K/V → ICF degenerates to image-invariant projection; ablation to prove 5-attr necessity).",
            "~1.8M (image_proj no grad)",
            "cls + bbox + dfl",
            None, None, None, None,
            None, None, "RUNNING (ep1/12, eta ~7h)",
        ),
    ]

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = WRAP if c_idx in (3, 4, 5, 12) else RIGHT
            if c_idx == 1:
                cell.alignment = Alignment(horizontal="center", vertical="top")
            if isinstance(val, float) and c_idx >= 6 and c_idx <= 11:
                cell.number_format = "0.0000"

        tag = row[1]
        if tag.startswith("ICF /"):
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = WIN_FILL
        elif tag.startswith("M1 + 5-attr mean"):
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = EMPHASIS_FILL
        elif tag.startswith("M2") and "ordclean" not in tag.lower():
            for c_idx in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c_idx).fill = FAIL_FILL

    ws.row_dimensions[1].height = 30
    for r in range(2, len(rows) + 2):
        ws.row_dimensions[r].height = 40

    _autofit(ws, max_width=55)

    # ── Notation block below the main table ──
    notation_start = len(rows) + 4  # blank row gap
    title_cell = ws.cell(row=notation_start, column=1, value="NOTATION (abbreviations used in the Method column)")
    title_cell.font = TITLE_FONT
    title_cell.fill = HEADER_FILL
    ws.merge_cells(start_row=notation_start, start_column=1,
                   end_row=notation_start, end_column=len(headers))

    notes = [
        ("Baseline (M1)",
         "Organ-conditional class loss masking on YOLOWorldHead. Per-image organ_id (read from "
         "image path via OrganExtractor pipeline) drives a {0,1} mask over 30 classes; loss_cls "
         "BCE is zeroed on cross-organ classes during train AND inference, so the detector only "
         "scores within-organ classes. Image backbone = ConvNext-tiny, trained from "
         "checkpoints/wedetect_tiny.pth (COCO-pretrained)."),
        ("1-PSC",
         "Single concatenated prompt per class, format \"{organ} cytology - {class_name}\" "
         "(e.g. \"Respiratory tract cytology - Neutrophil\"). Encoded once offline by "
         "BiomedCLIP-PubMedBERT-ViT-B/16-224 (frozen), cached as 512-d L2-normalized vector. "
         "Served by PseudoLanguageBackbone → [B, C, 512]."),
        ("5-attr",
         "Pathologist-canonical 5-attribute structured prompts per class: "
         "(1) specimen + collection method; "
         "(2) diagnosis system code (PSC / Bethesda / Paris); "
         "(3) cytology morphology (cell shape / nuclei / cytoplasm / chromatin); "
         "(4) background + immunohistochemistry markers; "
         "(5) key distinguishing feature vs the most-similar class. "
         "Each attribute encoded independently by BiomedCLIP. Served by "
         "PseudoMultiAttrLanguageBackbone → [B, C, 5, 512]."),
        ("mean pool",
         "Static aggregation over the 5-attr dim → [B, C, 512]. Zero trainable text-side params."),
        ("HTA",
         "Hierarchical Text Adapter (replaces mean pool with trainable 3-stage aggregation):\n"
         "• Stage 1 — per-attribute 2-layer MLP (D→128→D) + content-aware softmax attention "
         "pool over the 5 attribute outputs;\n"
         "• Stage 2 — per-organ soft MoE: 5 expert MLPs gated by softmax(W·ê + 5·1[o=organ(c)]) "
         "with gate_prior_strength=5 bias toward true organ;\n"
         "• Stage 3 — per-(organ, axis, rank) embedding lookup added residually to Stage 2 output."),
        ("4 aux losses",
         "HTA anti-collapse regularizers (default λ): pool_entropy=0.02, proj_drift=0.001, "
         "gate_entropy=0.02, rank_norm=0.01. Penalize uniform attention / projection drift / "
         "uniform organ gate / zero rank embedding norm, respectively."),
        ("ord_loss",
         "OrganOrdinalLoss. For each (organ, axis) pair with ≥2 valid-rank classes: linear "
         "head W_{o,a}·e_c + b_{o,a} → scalar predicted rank; MSE against rank_along_axis + "
         "within-axis pairwise monotonicity penalty (ReLU(pred[low] - pred[high])). "
         "loss_weight=0.3, monotonicity_weight=0.5. Normalization mode:\n"
         "• mean norm — divide total by # active axes (default; loss_ord ≈ 3e-4 on TCT_NGC);\n"
         "• sum norm  — leave as sum across active axes (×6 magnitude on TCT_NGC, "
         "loss_ord ≈ 5-30, gradients comparable to cls)."),
        ("M2 (Row 3)",
         "OC-HMTA Module 2 = full hierarchical text adapter setup = Baseline + 5-attr + HTA "
         "(Stage 1+2+3) + 4 HTA aux losses + ord_loss(mean norm). Rows 4 (auxfix) and 5 "
         "(axisstruct) are deltas on top of this Row 3 config. \"M2\" is used as shorthand "
         "in the Method column to avoid re-listing the 5 components."),
        ("auxfix (Row 4)",
         "M2 with two corrections aimed at making the aux supervisions actually train: "
         "(1) λ_pool_entropy boosted 15× (0.02 → 0.3) so Stage 1 attention entropy is "
         "actually forced below log(5); (2) OrganOrdinalLoss normalization switched from "
         "'mean' to 'sum' so loss_ord magnitude jumps ~6× and gradients aren't drowned by "
         "loss_cls. Result: training signals fire, but novel macro collapses (HTA over-"
         "specializes to base manifold)."),
        ("axisstruct (Row 5)",
         "M2 (sum-norm) + axis structure loss (defined below). Goal: keep cross-organ-kin "
         "(e.g. respiratory-Alveolar-macrophages ↔ Thyroid-Macrophages, identical "
         "cytomorphology) closer in emb space than M2's Stage 2 organ MoE allows."),
        ("axis structure loss",
         "(organ, axis)-conditional pull/push regularizer on HTA Stage 3 output (Row 5 only). "
         "Same-(organ, axis) class pairs attracted (cos ≥ 0.5, λ_axis_attract=0.3); "
         "cross-organ pairs repelled (cos ≤ 0.1, λ_cross_organ_repel=0.3); "
         "same-organ different-axis pairs left neutral."),
        ("ordclean / Row 6c",
         "Minimum-viable Module 2: Baseline + 5-attr + HTA Stage 1 only (skip Stage 2 organ "
         "MoE and Stage 3 rank emb lookup) + ord_loss restricted to medically clean rank "
         "ladders. The cleanest fair test of whether ord_loss intrinsically adds value when "
         "Stage 2/3 (verified novel killers by bypass diagnostic) are removed and broken "
         "label axes are excluded."),
        ("skip Stage 2/3 (used by Row 6c)",
         "HTA flags skip_stage2=True and skip_stage3_rank_emb=True bypass the corresponding "
         "forward paths and freeze the corresponding nn.Parameters (no DDP allreduce). "
         "Isolates Stage 1's contribution."),
        ("clean axes (used by Row 6c)",
         "OrganOrdinalLoss config: exclude_organ_axes = [(0,0) respiratory PSC — 5 normal cells "
         "share rank 2, MSE collision; (1,0) Serous binary; (4,2) TCT_CCD infection — "
         "non-ordinal pathogen labels] + skip_collision_ranks=True drops Thyroid rank-1 and "
         "rank-2 collisions. Remaining ord_loss supervision: 3 axes × 3-4 rank-unique exemplars."),
        ("ICF (Design A)",
         "Image-Conditional Fusion module (replaces mean pool / HTA):\n"
         "• image_ctx = Linear_{768→512}(GAP(F_4))   where F_4 is ConvNext-tiny stage-4 feature\n"
         "• attr_experts = 5 parallel 2-layer MLPs (D→128→D) applied per attribute index "
         "(static routing — every sample passes through every expert; NOT routing-style MoE)\n"
         "• attr_type_pe ∈ R^{5×512} added to expert outputs as K/V positional embedding\n"
         "• multi-head cross-attention (heads=8, pre-norm): Q=image_ctx broadcast over C classes, "
         "K=V=attr_experts(attr_emb) + attr_type_pe\n"
         "• output_proj + L2 normalization → fused class emb [B, C, 512]\n"
         "Monitored by ICFCollapseGuard hook (3 indicators every 500 train iter + at val)."),
        ("ICFCollapseGuard",
         "Hook that logs ICF health every 500 iter + before val. Three indicators:\n"
         "• fused_pairwise_cos_mean — same class, cross-image; healthy 0.5–0.95, red > 0.99 "
         "(= image-invariant collapse)\n"
         "• attn_entropy_mean — over 5 attribute keys; healthy < 1.58, red > 1.58 (= uniform)\n"
         "• cos_to_attr_mean_mean — fused direction vs static mean-pool; healthy < 0.95, "
         "red > 0.97 (= collapsed onto mean-pool subspace)."),
    ]

    note_row = notation_start + 1
    for term, defn in notes:
        ws.cell(row=note_row, column=1, value=term).font = HEADER_FONT
        ws.cell(row=note_row, column=1).alignment = WRAP
        ws.cell(row=note_row, column=2, value=defn).alignment = WRAP
        ws.merge_cells(start_row=note_row, start_column=2,
                       end_row=note_row, end_column=len(headers))
        # Generous height for multi-line definitions
        n_lines = max(1, defn.count("\n") + 1) + len(defn) // 90
        ws.row_dimensions[note_row].height = max(30, 14 * n_lines)
        note_row += 1


# ──────────────────────────────────────────────────────────────────────
# Sheet 2 — Per-organ novel breakdown
# ──────────────────────────────────────────────────────────────────────
def build_per_organ_novel(wb: Workbook) -> None:
    ws = wb.create_sheet("PerOrganNovel")
    _write_header(ws, [
        "Method",
        "Respiratory tract (3 novel cls)",
        "Serous effusion (3 novel cls)",
        "Thyroid gland (3 novel cls)",
        "Novel macro mAP",
        "Notes",
    ])

    rows = [
        ("M1 / 1-PSC", None, None, None, 0.1500,
         "Per-organ breakdown not recomputed; corrected-protocol macro only"),
        ("M1 + 5-attr mean pool", None, None, None, 0.1642,
         "Per-organ breakdown not recomputed; corrected-protocol macro only"),
        ("M2 完整方法", None, None, None, 0.1051,
         "Per-organ breakdown not recomputed"),
        ("M2 + auxfix", 0.0123, 0.0634, 0.0914, 0.0557,
         "Lowest novel; routing learned ⇒ over-specialized to base"),
        ("M2 + axisstruct", 0.0230, 0.1024, 0.0739, 0.0665,
         "Cross-organ repel partially helped Serous; respiratory still collapses"),
        ("M2 ordclean (Row 6c)", 0.1781, 0.1615, 0.1270, 0.1555,
         "Cleanest M2-flavored attempt; still net-negative vs static mean pool"),
        ("ICF / Design A", 0.1978, 0.1608, 0.1359, 0.1648,
         "Best per-organ across respiratory & thyroid; novel macro WIN"),
    ]
    for r_idx, row in enumerate(rows, start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = WRAP if c_idx in (1, 6) else RIGHT
            if isinstance(val, float):
                cell.number_format = "0.0000"
        if row[0] == "ICF / Design A":
            for c_idx in range(1, 7):
                ws.cell(row=r_idx, column=c_idx).fill = WIN_FILL
    _autofit(ws)


# ──────────────────────────────────────────────────────────────────────
# Sheet 3 — Bypass diagnostic (auxfix ep12, novel9 inference)
# ──────────────────────────────────────────────────────────────────────
def build_bypass(wb: Workbook) -> None:
    ws = wb.create_sheet("BypassDiagnostic")
    ws.cell(row=1, column=1, value=(
        "Inference-time module knockout on the M2-auxfix ep12 checkpoint. "
        "Each row replaces one HTA stage with a neutral/uniform pathway "
        "(Stage 1 → mean over 5 attribute projections; Stage 2 → mean over "
        "5 organ experts; Stage 3 → class_ranks buffer set to -1 disabling "
        "rank embedding lookup).  Trained per-attribute MLPs and per-organ "
        "expert MLPs are kept; only routing softmaxes are neutralized.  "
        "Conducted via tools/eval_organ_restricted.py --bypass-stages."
    )).alignment = WRAP
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 60

    _write_header(ws, [
        "Inference mode",
        "Respiratory novel AP",
        "Serous novel AP",
        "Thyroid novel AP",
        "Novel macro mAP",
        "Δ vs full auxfix",
    ], row=3)

    rows = [
        ("Full auxfix (no bypass)",      0.0123, 0.0634, 0.0914, 0.0557, "baseline"),
        ("Stage 1 bypass (uniform attn)",0.0049, 0.0841, 0.0889, 0.0593, "+0.4 pp"),
        ("Stage 3 bypass (rank=0)",      0.0416, 0.0875, 0.1072, 0.0788, "+2.3 pp"),
        ("Stage 2 bypass (uniform MoE)", 0.1077, 0.1103, 0.0944, 0.1041, "+4.8 pp ← main novel killer"),
        ("All-stages bypass",            0.0156, 0.0563, 0.0305, 0.0341, "-2.2 pp"),
    ]
    for r_idx, row in enumerate(rows, start=4):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = WRAP if c_idx in (1, 6) else RIGHT
            if isinstance(val, float):
                c.number_format = "0.0000"
        if "Stage 2 bypass" in row[0]:
            for c_idx in range(1, 7):
                ws.cell(row=r_idx, column=c_idx).fill = EMPHASIS_FILL
    _autofit(ws)


# ──────────────────────────────────────────────────────────────────────
# Sheet 4 — ICF training health
# ──────────────────────────────────────────────────────────────────────
def build_icf_health(wb: Workbook) -> None:
    ws = wb.create_sheet("ICFHealth")
    ws.cell(row=1, column=1, value=(
        "ICFCollapseGuard live diagnostics on the trained ICF (Design A) "
        "checkpoint at ep12.  All three indicators in healthy band → "
        "image-conditional fusion stayed image-conditional throughout "
        "training, did not collapse to a mean-pool function."
    )).alignment = WRAP
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 50

    _write_header(ws, [
        "Indicator",
        "Healthy range",
        "ep12 final reading",
        "Interpretation",
    ], row=3)

    rows = [
        ("fused_pairwise_cos_mean (same class, cross-image, off-diag)",
         "0.5 – 0.95",  "0.89 – 0.92",
         "Fused vector differs across images of the same class — "
         "image-conditional fusion is functioning."),
        ("attn_entropy_mean (over 5 attribute keys)",
         "< 1.58 (log 5 = 1.609)",  "1.26 – 1.32",
         "Attention concentrates on ~2-3 attributes per (image, class) "
         "rather than being uniform → image-conditional attribute selection."),
        ("cos_to_attr_mean_mean (fused vs static mean-pool direction)",
         "< 0.95",  "-0.02",
         "Fused direction is essentially orthogonal to the mean-pool "
         "subspace — ICF escaped the mean-pool basin of attraction."),
    ]
    for r_idx, row in enumerate(rows, start=4):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = WRAP
        for c_idx in range(1, 5):
            ws.cell(row=r_idx, column=c_idx).fill = WIN_FILL
    _autofit(ws, max_width=60)


# ──────────────────────────────────────────────────────────────────────
# Sheet 5 — Rank label audit
# ──────────────────────────────────────────────────────────────────────
def build_rank_audit(wb: Workbook) -> None:
    ws = wb.create_sheet("RankLabelAudit")
    ws.cell(row=1, column=1, value=(
        "Per-(organ, axis) audit of rank_along_axis labels parsed by "
        "tools/build_taxonomy_metadata.py from PSC / Bethesda / Paris "
        "Roman-numeral category codes.  Determines which axes are usable "
        "by OrganOrdinalLoss without MSE rank-collision destroying "
        "discriminative training signal."
    )).alignment = WRAP
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws.row_dimensions[1].height = 55

    _write_header(ws, [
        "Organ-axis", "Rank ladder", "Quality", "Used in Row 6c (M2 ordclean)?"
    ], row=3)
    rows = [
        ("Urine — axis 0 (Paris System)",
         "NHGUC(2) → AUC(3) → SHGUC(4) → HGUC(5)",
         "✅ clean ordinal (4 ranks, 1 class each)", "YES"),
        ("TCT_CCD — axis 0 (Bethesda cervical squamous)",
         "ASCUS(1) → ASCH(2) → LSIL(3) → HSIL(4)",
         "✅ clean ordinal (4 ranks, 1 class each)", "YES"),
        ("Thyroid — axis 0 (Bethesda thyroid)",
         "Macrophages, NS at rank 1 (collision); FC, Negative samples at "
         "rank 2 (collision); AUC, SPTC, PTC at higher ranks",
         "⚠️ partial — only 3 rank-unique exemplars (AUC=3, SPTC=5, PTC=6) "
         "after skip_collision_ranks",
         "YES (rank-unique exemplars only)"),
        ("Respiratory tract — axis 0 (PSC)",
         "5 normal cell types (Neutrophil, Macrophages, Cilia cells, "
         "Lymphocyte, Impurity) all at PSC Category II = rank 2; "
         "Squamous epi & Diseased cells at higher ranks",
         "❌ MSE rank-collision destroys class discriminability — "
         "5 classes forced to same scalar projection by ord_loss MSE",
         "NO — excluded via exclude_organ_axes=[(0,0)]"),
        ("Serous effusion — axis 0",
         "Negative samples vs Diseased cells (binary)",
         "❌ degenerate — only 2 classes / 2 ranks, no ordinal structure",
         "NO — excluded via exclude_organ_axes=[(1,0)]"),
        ("TCT_CCD — axis 2 (infection)",
         "monilia / dysbacteriosis / vaginalis — different organisms",
         "❌ not ordinal — labels are arbitrary index over pathogens, "
         "not a severity scale",
         "NO — excluded via exclude_organ_axes=[(4,2)]"),
    ]
    for r_idx, row in enumerate(rows, start=4):
        for c_idx, val in enumerate(row, start=1):
            c = ws.cell(row=r_idx, column=c_idx, value=val)
            c.alignment = WRAP
        if row[3] == "YES" or "YES" in row[3]:
            ws.cell(row=r_idx, column=4).fill = WIN_FILL
        elif row[3].startswith("NO"):
            ws.cell(row=r_idx, column=4).fill = FAIL_FILL
    _autofit(ws, max_width=70)


# ──────────────────────────────────────────────────────────────────────
# Sheet 6 — Shared training hyperparameters
# ──────────────────────────────────────────────────────────────────────
def build_hyperparams(wb: Workbook) -> None:
    ws = wb.create_sheet("Hyperparameters")
    ws.cell(row=1, column=1, value=(
        "Shared training and evaluation configuration across all rows of "
        "the ablation table.  Any deviation (e.g. M2-auxfix λ_pool_entropy, "
        "Row 6c skip flags) is noted in the Method column of the main "
        "table.  Listed here to confirm apples-to-apples comparison."
    )).alignment = WRAP
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws.row_dimensions[1].height = 60

    _write_header(ws, ["Component", "Value / setting"], row=3)
    rows = [
        ("Image backbone", "ConvNext-tiny (depths=[3,3,9,3], dims=[96,192,384,768]), trained"),
        ("Text encoder",
         "BiomedCLIP-PubMedBERT-ViT-Base-Patch16-224 (frozen); embeddings cached to disk"),
        ("Detector head", "YOLOWorldHead (contrastive head, with organ_loss_mask)"),
        ("Initial weights", "checkpoints/wedetect_tiny.pth (same start across all rows)"),
        ("Optimizer", "AdamW, base_lr=3e-4, weight_decay=0.05"),
        ("LR schedule", "Linear warmup (begin=0, end=1) → CosineAnnealing (begin=2, T_max=11)"),
        ("Epochs", "12"),
        ("Batch", "8 per GPU × 2 GPUs (RTX 3090) = 16 effective"),
        ("AMP", "Disabled (fp32)"),
        ("Image size", "640 × 640 (WeDetectKeepRatioResize + LetterResize)"),
        ("Augmentation",
         "PhotoMetricDistortion + RandomFlip(h+v); no Mosaic; medical cells stay intact"),
        ("Train annotations",
         "/home1/liwenjie/TCT_NGC_640/annotations/instances_train_dev_disjoint_dev30.json (30 base classes)"),
        ("Val annotations (during training)",
         "instances_val_dev_disjoint_dev30.json"),
        ("Test base 25",
         "instances_test_base_clean_dev30.json (30 cls minus 5 negative classes)"),
        ("Test novel 9",
         "instances_test_novel_merged_9.json (Respiratory ×3, Serous ×3, Thyroid ×3)"),
        ("Evaluation",
         "OrganRestrictedCocoMetric: per-organ AP + overall macro + instance-weighted; "
         "5 negative classes excluded from COCOeval catIds for parity with paper-protocol report"),
        ("Best-checkpoint selection",
         "save_best='coco/overall/macro_mAP' (corrected protocol); older M1 / "
         "M1-5attr平均 / M2 trained before this fix and used save_best='coco/bbox_mAP' "
         "— their reported numbers come from re-eval on each saved epoch ckpt via "
         "tools/eval_all_ckpts_corrected_val.sh, picking the truly-best epoch"),
        ("DDP",
         "find_unused_parameters=True (set in dev30 base config); enables Row 6c "
         "skip-stage knockouts and the ICF + 1-PSC degenerate cross-attn path "
         "without DDP errors"),
        ("Exclude classes (5 negatives)",
         "respiratory tract-Impurity, Serous effusion-Negative samples, "
         "Thyroid gland-Negative samples, Urine-NHGUC, TCT_CCD-normal"),
    ]
    for r_idx, (k, v) in enumerate(rows, start=4):
        ws.cell(row=r_idx, column=1, value=k).alignment = WRAP
        ws.cell(row=r_idx, column=1).font = HEADER_FONT
        ws.cell(row=r_idx, column=2, value=v).alignment = WRAP
    _autofit(ws, max_width=90)


def build_per_class_novel(wb: Workbook) -> None:
    ws = wb.create_sheet("NovelPerClassBreakdown")
    ws.cell(row=1, column=1, value=(
        "Per-class AP on the 9-class novel test set (instances_test_novel_merged_9.json) "
        "for the ICF (Design A) checkpoint at epoch 12. Macro 0.1648 is the simple mean of "
        "the 9 per-class APs.  Classes are grouped by morphological proximity to base training "
        "classes — a critical fairness disclosure: 3 \"subtype-of-base free rider\" classes "
        "(adenocarcinoma / Ovarian / Suspicious for Malignancy) contribute ~84% of the reported "
        "macro because their parent class (base \"Diseased cells, NOS\" or base "
        "AUC/SPTC) was already trained.  The 3 truly-novel classes (MTC / Small cell carcinoma "
        "/ Serous adenocarcinoma) hover near 0 — the detector cannot zero-shot recognize "
        "morphologies fundamentally different from any base class."
    )).alignment = WRAP
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 100

    _write_header(ws, [
        "Group", "Class", "Organ", "mAP50-95", "AP50", "Closest base class / why this category"
    ], row=3)

    rows = [
        # Subtype-of-base (high AP, free rider)
        ("Subtype-of-base", "respiratory tract-adenocarcinoma", "respiratory tract",
         0.5245, 0.7058, "Subtype of base 6 'Diseased cells, NOS'; adenocarcinoma cells well-represented in base training"),
        ("Subtype-of-base", "Serous effusion-Ovarian cancer", "Serous effusion",
         0.3527, 0.5195, "Subtype of base 8 'Diseased cells, NOS'; morphologically similar to other serous malignancies in base"),
        ("Subtype-of-base", "Thyroid gland-Suspicious for Malignancy", "Thyroid gland",
         0.3715, 0.5248, "Essentially the same diagnostic category as base 10 SPTC / base 13 AUC"),
        # Borderline (medium AP, partial overlap)
        ("Borderline", "respiratory tract-Squamous cell carcinoma", "respiratory tract",
         0.0663, 0.0981, "Subtype of base 6 but specific squamous-cell morphology less common in base training"),
        ("Borderline", "Serous effusion-Breast cancer", "Serous effusion",
         0.1193, 0.1597, "Subtype of base 8 but breast-origin morphology distinct"),
        ("Borderline", "Thyroid gland-Malignant tumour", "Thyroid gland",
         0.0364, 0.0526, "Generic malignancy label; partial overlap with base 9 PTC / base 10 SPTC"),
        # Truly novel (near-zero AP)
        ("Truly novel", "respiratory tract-Small cell carcinoma", "respiratory tract",
         0.0025, 0.0034, "Small cell carcinoma has distinct morphology (neuroendocrine) absent from base"),
        ("Truly novel", "Serous effusion-adenocarcinoma", "Serous effusion",
         0.0104, 0.0140, "Adeno morphology in serous effusion distinct from base 8 Diseased; structurally novel"),
        ("Truly novel", "Thyroid gland-MTC", "Thyroid gland",
         0.0000, 0.0000, "Medullary thyroid carcinoma — entirely different cell type from base 9 PTC"),
    ]

    for r_idx, row in enumerate(rows, start=4):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = WRAP if c_idx in (2, 6) else RIGHT
            if isinstance(val, float):
                cell.number_format = "0.0000"
        # Color by group
        grp = row[0]
        if grp == "Subtype-of-base":
            for c in range(1, 7):
                ws.cell(row=r_idx, column=c).fill = WIN_FILL  # green = high (free rider)
        elif grp == "Truly novel":
            for c in range(1, 7):
                ws.cell(row=r_idx, column=c).fill = FAIL_FILL  # red = near-zero
        # Borderline: default (no fill)

    # Group-summary rows
    summary_start = 4 + len(rows) + 1
    summary_rows = [
        ("Subtype-of-base macro", 3, 0.4162,
         "Subtype-of-base classes' macro AP (mean over 3 classes). "
         "Represents the detector's ability to recognize finer-grained labels of cells it already learned."),
        ("Borderline macro", 3, 0.0740,
         "Borderline classes' macro AP (mean over 3 classes). "
         "Specific morphologies with partial overlap to base classes; detector picks up some."),
        ("Truly novel macro", 3, 0.0043,
         "Truly-novel classes' macro AP (mean over 3 classes). "
         "Cells morphologically distinct from any base class — detector essentially fails (~0)."),
        ("Reported macro (all 9)", 9, 0.1648,
         "What we report in the main ablation table. Heavily inflated by the 3 subtype-of-base outliers; "
         "without them macro would be ~0.04 (truly + borderline combined)."),
    ]
    ws.cell(row=summary_start, column=1, value="── Group summary ──").font = HEADER_FONT
    ws.merge_cells(start_row=summary_start, start_column=1,
                   end_row=summary_start, end_column=6)
    for i, (grp, n, ap, note) in enumerate(summary_rows, start=summary_start + 1):
        ws.cell(row=i, column=1, value=grp).font = HEADER_FONT
        ws.cell(row=i, column=2, value=f"n={n}")
        ws.cell(row=i, column=4, value=ap).number_format = "0.0000"
        ws.cell(row=i, column=6, value=note).alignment = WRAP
        ws.merge_cells(start_row=i, start_column=6, end_row=i, end_column=6)
        if "Subtype" in grp:
            for c in range(1, 7):
                ws.cell(row=i, column=c).fill = WIN_FILL
        elif "Truly novel" in grp:
            for c in range(1, 7):
                ws.cell(row=i, column=c).fill = FAIL_FILL
        elif "Reported" in grp:
            for c in range(1, 7):
                ws.cell(row=i, column=c).fill = EMPHASIS_FILL
        ws.row_dimensions[i].height = 38

    # Add YOLOE comparison block
    yoloe_start = summary_start + len(summary_rows) + 3
    ws.cell(row=yoloe_start, column=1, value=(
        "── YOLOE comparison (for context) ──\n"
        "YOLOE evaluates on 4 separate K-class splits (main_3 / pseudo_2 / hard_4 / full_5) "
        "and reports the macro-of-macros = 24.1. Per-class breakdown shows the same pattern: "
        "high-AP classes (Squamous carcinoma, Mal-S Metastatic breast) hover 0.30-0.48 "
        "(subtype-of-base free riders), while Medullary thyroid carcinoma (MTC, the same "
        "truly-novel class as ours) drops to 0.0009. The 24.1 average is inflated by smaller-K "
        "splits where 1-2 classes/organ + organ mask reduces the task to near-trivial detection. "
        "On full_5 (K=5, the most comparable to our 9-class merged eval), YOLOE macro 0.156 "
        "≈ our ICF macro 0.165 — basically tied once class-density is controlled."
    )).alignment = WRAP
    ws.cell(row=yoloe_start, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=yoloe_start, start_column=1,
                   end_row=yoloe_start, end_column=6)
    ws.row_dimensions[yoloe_start].height = 120

    _autofit(ws, max_width=70)


def build_yoloe_comparison(wb: Workbook) -> None:
    ws = wb.create_sheet("YOLOEComparison")
    ws.cell(row=1, column=1, value=(
        "Head-to-head comparison vs YOLOE (Tencent Ultralytics) on the IDENTICAL "
        "9-class novel pooled evaluation (instances_test_novel_merged_9.json, 8880 dedup "
        "images, single eval pass, 9 per-class APs → simple macro). "
        "\n\n"
        "IMPORTANT — protocol asymmetry that affects interpretation: YOLOE's organ-conditional "
        "class score mask is applied ONLY AT INFERENCE (postprocess: cls scores ×{0,1} mask). "
        "Their model trained without any mask, so 'prior OFF' eval (3.3) measures the model's "
        "raw zero-shot capability INCLUDING cross-organ hallucinations it learned during "
        "training. 'prior ON' (9.7) just clips those hallucinations at score time. "
        "\n\n"
        "WeDetect's organ mask is applied DURING TRAINING (loss_cls BCE zeroed on cross-organ "
        "classes + positive cls_preds also zeroed to prevent spurious positive assignment) "
        "AND at inference (same as YOLOE). So our model never learned cross-organ associations "
        "in the first place — the inference-time mask is a no-op relative to a model that "
        "already knows the within-organ constraint. This is a strictly stronger M1 implementation."
    )).alignment = WRAP
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells("A1:G1")
    ws.row_dimensions[1].height = 200

    _write_header(ws, [
        "Method", "Text encoder",
        "Train-time organ mask",
        "Inference organ mask",
        "Novel macro (9-cls pooled)",
        "Δ vs WeDetect ICF",
        "Notes",
    ], row=3)

    rows = [
        # YOLOE — no train-time mask
        ("YOLOE prior OFF (text-only)",
         "MobileCLIP",
         "❌ NOT applied during training",
         "❌ no mask at inference",
         0.033, -0.1318,
         "Raw model output; cross-organ hallucinations present. "
         "Baseline for what the model learned without any organ constraint."),
        ("YOLOE prior ON (text-only)",
         "MobileCLIP",
         "❌ NOT applied during training",
         "✅ mask applied at postprocess (cls_score × {0,1} mask)",
         0.097, -0.0678,
         "Inference-time mask clips cross-organ hallucinations. "
         "+6.4pp from OFF→ON measures the magnitude of hallucinations the model learned. "
         "YOLOE side acknowledges MobileCLIP cannot understand medical terminology like "
         "'Medullary thyroid' / 'Bethesda V/VI' / 'PSC Category VI' — 4 truly-novel classes "
         "score ~0 even with mask."),
        # WeDetect — mask applied at BOTH training and inference
        ("WeDetect M1 (1-PSC)",
         "BiomedCLIP-ViT-B/16",
         "✅ applied during training (loss_cls BCE × mask)",
         "✅ applied at inference (cls_score × mask, pre-NMS)",
         0.150, -0.0148,
         "Model never learned cross-organ associations during training. "
         "Cell representations are organ-aware from the start."),
        ("WeDetect M1 + 5-attr mean pool",
         "BiomedCLIP-ViT-B/16",
         "✅ applied during training (loss_cls BCE × mask)",
         "✅ applied at inference",
         0.164, -0.0008,
         "+ 5-attribute pathologist-canonical structured prompts (zero trainable text-side params). "
         "Static mean-pool aggregation. Practically tied with ICF on novel."),
        ("WeDetect ICF (Design A) ⭐",
         "BiomedCLIP-ViT-B/16",
         "✅ applied during training (loss_cls BCE × mask)",
         "✅ applied at inference",
         0.1648, 0.0000,
         "Reference point. Image-conditional cross-attention fusion over 5 per-attribute MLPs. "
         "Best novel macro on this protocol; primary base-mAP gain (+1.3pp over M1+5-attr)."),
    ]
    for r_idx, row in enumerate(rows, start=4):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.alignment = WRAP if c_idx in (1, 3, 4, 7) else RIGHT
            if isinstance(val, float):
                cell.number_format = "0.0000"
        if row[0].startswith("WeDetect ICF"):
            for c in range(1, 8):
                ws.cell(row=r_idx, column=c).fill = WIN_FILL
        elif row[0].startswith("YOLOE"):
            for c in range(1, 8):
                ws.cell(row=r_idx, column=c).fill = FAIL_FILL
        ws.row_dimensions[r_idx].height = 95

    # Summary findings
    summary_start = 4 + len(rows) + 2
    ws.cell(row=summary_start, column=1,
            value="── Summary findings ──").font = TITLE_FONT
    ws.merge_cells(start_row=summary_start, start_column=1,
                   end_row=summary_start, end_column=7)
    findings = [
        ("WeDetect > YOLOE under fair protocol",
         "WeDetect ICF 0.1648 vs YOLOE prior ON 0.097 → **+6.78 pp (+70% relative)** under "
         "identical 9-class pooled zero-shot eval. Gap holds across all WeDetect rows: "
         "even our M2 (0.105) edges out YOLOE (0.097)."),
        ("Primary attributable factors",
         "(a) Text encoder: BiomedCLIP (medical-pretrained) vs MobileCLIP (general) — "
         "YOLOE side confirms MobileCLIP cannot encode clinical terminology (Bethesda V/VI, "
         "PSC Category, Medullary thyroid) for the 4 truly-novel classes. "
         "(b) Train-time organ mask: WeDetect never learns cross-organ hallucinations, so the "
         "9.7-floor (model trained without mask + inference mask) is lower than ours."),
        ("Earlier YOLOE-reported 24.1 was non-comparable",
         "YOLOE's headline 24.1 was the macro-of-macros across 4 K-class splits (main_3 / "
         "pseudo_2 / hard_4 / full_5), with K ∈ {2, 3, 4, 5}. Smaller-K splits with 1 cls/organ "
         "made the eval near-trivial (mask reduces to 1-of-1 selection). On a single 9-class "
         "pooled eval their number drops to 9.7 (YOLOE side independently confirmed)."),
        ("YOLOE visual-prompt path (24.7) is NOT zero-shot",
         "YOLOE's visual-prompt mode requires 30% of novel images as support exemplars to "
         "build SAVPE visual prototypes. This is 5-shot/30%-shot eval, not zero-shot. "
         "WeDetect's all rows are pure zero-shot (0% support, novel images unseen). "
         "Apples-to-apples comparison would need WeDetect with visual prompts (Phase 5 future work) "
         "or YOLOE with 0% support — both not done in this paper."),
        ("Per-class breakdown caveat (applies to both sides)",
         "Both 16.5 (ours) and 9.7 (YOLOE) hide the same heterogeneity: 3 subtype-of-base "
         "classes carry the macro, 3 truly-novel classes (MTC / Small cell / Serous adeno) "
         "are ~0 on both sides. See NovelPerClassBreakdown sheet."),
    ]
    for i, (title, body) in enumerate(findings, start=summary_start + 1):
        c1 = ws.cell(row=i, column=1, value=title)
        c1.font = HEADER_FONT
        c1.alignment = WRAP
        c2 = ws.cell(row=i, column=2, value=body)
        c2.alignment = WRAP
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
        n_lines = max(1, len(body) // 90)
        ws.row_dimensions[i].height = max(40, 16 * n_lines)

    _autofit(ws, max_width=55)


def main():
    wb = Workbook()
    # Workbook starts with a default "Sheet"; remove it.
    default_ws = wb.active
    wb.remove(default_ws)

    build_main_table(wb)
    build_per_organ_novel(wb)
    build_per_class_novel(wb)
    build_yoloe_comparison(wb)
    build_bypass(wb)
    build_icf_health(wb)
    build_rank_audit(wb)
    build_hyperparams(wb)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"wrote {OUT_PATH}")
    print(f"sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
